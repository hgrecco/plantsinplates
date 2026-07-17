import json
import pathlib
import re
import tempfile
import threading
import unittest
from types import SimpleNamespace
from unittest import mock

from openpyxl import Workbook, load_workbook

from plantsinplates import analyze, workflow
from plantsinplates.measurement_config import MeasurementConfig


def make_plate(
    root: pathlib.Path, *, name: str = "plate_001", image_suffix: str = ".png"
) -> pathlib.Path:
    plate = root / name
    image = plate / "date_20260101" / "row_001" / f"fluo_1_1{image_suffix}"
    image.parent.mkdir(parents=True)
    image.write_bytes(b"not read by mocked measurement")
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["row", "col", "genotype", "20260101_length", "20260101_fluo"])
    sheet.append([1, 1, "control", 10.0, f"row_001/fluo_1_1{image_suffix}"])
    workbook.save(plate / "info.xlsx")
    return plate


class WorkflowTests(unittest.TestCase):
    def test_validation_checks_references_and_reports_unreferenced_images(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = pathlib.Path(temporary)
            plate = make_plate(root)
            extra = plate / "date_20260101" / "extra.png"
            extra.write_bytes(b"extra")
            result = workflow.validate_folder(plate)
            self.assertTrue(result.valid)
            self.assertEqual(result.image_count, 1)
            self.assertEqual(result.unreferenced_images, (extra.resolve(),))

            referenced = plate / "date_20260101" / "row_001" / "fluo_1_1.png"
            referenced.unlink()
            invalid = workflow.validate_folder(plate)
            self.assertFalse(invalid.valid)
            self.assertIn("missing referenced images", " ".join(invalid.details))

    def test_metadata_calibration_accepts_square_pixels_and_rejects_anisotropy(
        self,
    ) -> None:
        path = pathlib.Path("image.czi")
        with mock.patch.object(pathlib.Path, "is_file", return_value=True):
            with mock.patch.object(
                workflow.bioio_czi,
                "Reader",
                return_value=SimpleNamespace(
                    physical_pixel_sizes=SimpleNamespace(X=0.25, Y=0.251)
                ),
            ):
                reading = workflow.read_image_calibration(path)
            self.assertTrue(reading.valid)
            self.assertAlmostEqual(reading.um_per_pixel or 0, 0.2505)

            with mock.patch.object(
                workflow.bioio_czi,
                "Reader",
                return_value=SimpleNamespace(
                    physical_pixel_sizes=SimpleNamespace(X=0.25, Y=0.5)
                ),
            ):
                reading = workflow.read_image_calibration(path)
            self.assertEqual(reading.status, "non_square")

    def test_physical_settings_resolve_per_image(self) -> None:
        settings = MeasurementConfig(method="box", box_size=100, box_offset=0)
        request = workflow.AnalysisRequest(
            pathlib.Path("plate_001"),
            "plate",
            settings,
            "um",
            workflow.CalibrationSpec("metadata"),
        )
        first = analyze.resolve_pixel_config(request, 0.5)
        second = analyze.resolve_pixel_config(request, 0.25)
        self.assertEqual(first.box_size, 200)
        self.assertEqual(second.box_size, 400)

    def test_run_ids_and_discovery_require_a_manifest(self) -> None:
        run_id = workflow.generate_run_id()
        self.assertRegex(run_id, re.compile(r"^\d{8}-\d{6}-[0-9a-f]{8}$"))
        with tempfile.TemporaryDirectory() as temporary:
            folder = pathlib.Path(temporary) / "plate_001"
            folder.mkdir()
            _allocated_id, run_dir = workflow.create_run_directory(folder)
            (folder / "_output_legacy.pdf").write_bytes(b"legacy")
            manifest = {
                "run_id": run_dir.name.removeprefix("_output_"),
                "started_at": "2026-01-01T12:00:00+00:00",
                "status": "completed",
                "request": {
                    "settings": {"method": "box", "box_size": 100},
                    "settings_unit": "um",
                    "calibration": {"mode": "manual", "shared_um_per_pixel": 0.5},
                    "reuse_policy": "compatible",
                },
                "outputs": {"PDF summary": "summary.pdf"},
            }
            workflow.atomic_write_json(run_dir / "run.json", manifest)
            records = workflow.discover_runs(folder)
            self.assertEqual(len(records), 1)
            self.assertEqual(records[0].method, "box")

    def test_run_analysis_creates_isolated_outputs_and_reuses_measurement(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            plate = make_plate(pathlib.Path(temporary))
            request = workflow.AnalysisRequest(
                plate,
                "plate",
                MeasurementConfig(method="box", box_size=20, box_offset=0),
                "px",
                workflow.CalibrationSpec("pixels"),
                reuse_policy="compatible",
            )
            measured = {
                "tip_fg_mean": 10.0,
                "tip_bg_mean": 2.0,
                "tip_fg_std": 1.0,
                "tip_bg_std": 1.0,
                "tip_fg_count": 10,
                "tip_bg_count": 10,
                "tip_box": (20, 20),
                "tip_position": (10, 10),
                "full_fg_mean": 10.0,
                "full_bg_mean": 2.0,
            }
            events: list[workflow.ProgressEvent] = []
            with (
                mock.patch.object(
                    analyze.imsingleroot, "measure_image", return_value=measured
                ) as measure,
                mock.patch.object(analyze.visualize, "generate_analysis_details_page"),
                mock.patch.object(analyze.visualize, "generate_plateview"),
            ):
                first = analyze.run_analysis(request, progress=events.append)
            self.assertEqual(first.status, "completed")
            self.assertEqual(measure.call_count, 1)
            self.assertTrue((first.run_dir / "run.json").is_file())
            self.assertTrue((first.run_dir / "summary.xlsx").is_file())
            self.assertTrue((first.run_dir / "calibration_report.csv").is_file())
            self.assertTrue(any(event.stage == "measurement" for event in events))

            with (
                mock.patch.object(
                    analyze.imsingleroot, "measure_image"
                ) as second_measure,
                mock.patch.object(analyze.visualize, "generate_analysis_details_page"),
                mock.patch.object(analyze.visualize, "generate_plateview"),
            ):
                second = analyze.run_analysis(request)
            self.assertEqual(second.status, "completed", second.message)
            self.assertEqual(second.reused, 1)
            second_measure.assert_not_called()
            self.assertNotEqual(first.run_dir, second.run_dir)
            manifest = json.loads((second.run_dir / "run.json").read_text())
            self.assertEqual(manifest["stats"]["reused_measurements"], 1)

    def test_cancellation_leaves_a_manifest_backed_run(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            plate = make_plate(pathlib.Path(temporary))
            request = workflow.AnalysisRequest(
                plate,
                "plate",
                MeasurementConfig(),
                "px",
                workflow.CalibrationSpec("pixels"),
            )
            cancellation = threading.Event()
            cancellation.set()
            result = analyze.run_analysis(request, cancel_event=cancellation)
            self.assertEqual(result.status, "cancelled")
            manifest = json.loads((result.run_dir / "run.json").read_text())
            self.assertEqual(manifest["status"], "cancelled")

    def test_reuse_policies_control_measurement_and_preprocessing_reuse(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            plate = make_plate(pathlib.Path(temporary))
            measured = {"tip_fg_mean": 4.0, "tip_bg_mean": 1.0}
            for policy, expected_artifact_reuse in (
                ("preprocessing", True),
                ("none", False),
            ):
                request = workflow.AnalysisRequest(
                    plate,
                    "plate",
                    MeasurementConfig(method="box", box_size=20),
                    "px",
                    workflow.CalibrationSpec("pixels"),
                    reuse_policy=policy,
                )
                with (
                    mock.patch.object(
                        analyze.imsingleroot,
                        "measure_image",
                        return_value=measured,
                    ) as measure,
                    mock.patch.object(
                        analyze.visualize, "generate_analysis_details_page"
                    ),
                    mock.patch.object(analyze.visualize, "generate_plateview"),
                ):
                    result = analyze.run_analysis(request)
                self.assertEqual(result.status, "completed")
                self.assertEqual(
                    measure.call_args.kwargs["reuse_artifacts"],
                    expected_artifact_reuse,
                )

    def test_experiment_run_contains_merged_and_per_plate_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            experiment = pathlib.Path(temporary) / "experiment_001"
            make_plate(experiment, name="plate_001")
            make_plate(experiment, name="plate_002")
            request = workflow.AnalysisRequest(
                experiment,
                "experiment",
                MeasurementConfig(method="box", box_size=20),
                "px",
                workflow.CalibrationSpec("pixels"),
            )
            measured = {"tip_fg_mean": 4.0, "tip_bg_mean": 1.0}
            with (
                mock.patch.object(
                    analyze.imsingleroot, "measure_image", return_value=measured
                ),
                mock.patch.object(analyze.visualize, "generate_analysis_details_page"),
                mock.patch.object(analyze.visualize, "generate_plateview"),
                mock.patch.object(analyze.visualize, "generate_experimentview"),
            ):
                result = analyze.run_analysis(request)
            self.assertEqual(result.status, "completed")
            self.assertTrue((result.run_dir / "summary.xlsx").is_file())
            self.assertTrue(
                (result.run_dir / "plates" / "plate_001" / "summary.xlsx").is_file()
            )
            self.assertTrue(
                (result.run_dir / "plates" / "plate_002" / "summary.xlsx").is_file()
            )

    def test_metadata_mode_skips_only_images_with_unusable_calibration(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            plate = make_plate(pathlib.Path(temporary))
            second_image = plate / "date_20260101" / "row_002" / "fluo_2_1.png"
            second_image.parent.mkdir()
            second_image.write_bytes(b"second")
            workbook = load_workbook(plate / "info.xlsx")
            workbook.active.append([2, 1, "control", 11.0, "row_002/fluo_2_1.png"])
            workbook.save(plate / "info.xlsx")
            request = workflow.AnalysisRequest(
                plate,
                "plate",
                MeasurementConfig(method="box", box_size=20),
                "um",
                workflow.CalibrationSpec("metadata"),
            )

            def calibration(path: pathlib.Path) -> workflow.CalibrationReading:
                if "fluo_1_1" in path.name:
                    return workflow.CalibrationReading(
                        path,
                        "ok",
                        x_um_per_pixel=0.5,
                        y_um_per_pixel=0.5,
                        um_per_pixel=0.5,
                    )
                return workflow.CalibrationReading(
                    path, "invalid", message="metadata is missing"
                )

            with (
                mock.patch.object(
                    workflow, "read_image_calibration", side_effect=calibration
                ),
                mock.patch.object(
                    analyze.imsingleroot,
                    "measure_image",
                    return_value={"tip_fg_mean": 4.0, "tip_bg_mean": 1.0},
                ),
                mock.patch.object(analyze.visualize, "generate_analysis_details_page"),
                mock.patch.object(analyze.visualize, "generate_plateview"),
            ):
                result = analyze.run_analysis(request)
            self.assertEqual(result.status, "completed_with_warnings")
            self.assertEqual(result.recomputed, 1)
            self.assertEqual(result.skipped, 1)
            report = (result.run_dir / "calibration_report.csv").read_text()
            self.assertIn("metadata is missing", report)
            self.assertIn("skipped", report)


if __name__ == "__main__":
    unittest.main()
